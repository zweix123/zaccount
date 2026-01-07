import csv
from datetime import datetime
from typing import Callable

import openpyxl

from utils import check_categorys, get_data_file_path, load_ctg

ENTRY_KEYS = ["date", "type", "amount", "categorys", "tags", "desc"]

SUM_FACTOR: dict[str, float] = {
    "收入": 1,
    "支出": -1,
    "转入": 1,
    "转出": -1,
}
assert set(SUM_FACTOR.keys()) == set(
    load_ctg().keys()
), f"缺少或者多余的类别: {set(SUM_FACTOR.keys()) ^ set(load_ctg().keys())}"


class Entry:
    def __init__(
        self,
        date: datetime,
        type: str,
        amount: float,
        categorys: list[str],
        tags: list[str],
        desc: str,
    ) -> None:
        self.date = date
        self.type = type
        self.amount = amount
        self.categorys = categorys
        self.tags = tags
        self.desc = desc

    @classmethod
    def from_dict(cls, data: dict) -> "Entry":
        for key in ENTRY_KEYS:
            if key not in data:
                raise Exception(f"key {key} not in data")
        # date
        if data["date"] is None:
            raise Exception(f"date {data['date']} is None")
        if isinstance(data["date"], datetime):
            date_field = data["date"]
        else:
            try:
                date_field = datetime.strptime(data["date"], "%Y-%m-%d")
            except Exception as e:
                raise Exception(
                    f"date {data['date']} is not in the format %Y-%m-%d: {e}"
                )
        # type
        type_field = data["type"]
        if len(type_field) == 0:
            raise Exception(f"type {type_field} is empty")
        if type_field not in load_ctg():
            raise Exception(f"type {type_field} not in load_ctg()")
        # amount
        try:
            amount_field = float(data["amount"])
        except Exception:
            raise Exception(f"amount {data['amount']} is not a number")
        if amount_field <= 0:
            raise Exception(f"amount {amount_field} is not positive")
        # categorys
        try:
            categorys_field = data["categorys"].split(",")
        except Exception:
            raise Exception(
                f"categorys {data['categorys']} is not a list split by comma"
            )
        if not check_categorys(type_field, categorys_field):
            raise Exception(
                f"categorys {categorys_field} not in load_ctg()[{type_field}]"
            )
        # tags
        tags_field: list[str] = []
        if isinstance(data["tags"], str) and len(data["tags"]) != 0:
            try:
                tags_field = data["tags"].split(",")
            except Exception:
                raise Exception(f"tags {data['tags']} is not a list split by comma")
        # desc
        desc_field = data["desc"]

        return cls(
            date=date_field,
            type=type_field,
            amount=amount_field,
            categorys=categorys_field,
            tags=tags_field,
            desc=desc_field,
        )

    def to_dict(self) -> dict:
        transforms = {
            "date": lambda x: x.strftime("%Y-%m-%d"),
            "type": lambda x: x,
            "amount": lambda x: str(x),
            "categorys": lambda x: ",".join(x),
            "tags": lambda x: ",".join(x),
            "desc": lambda x: x,
        }
        return {key: transforms[key](getattr(self, key)) for key in ENTRY_KEYS}

    def __str__(self) -> str:
        return str(self.to_dict())

    def __repr__(self) -> str:
        return repr(self.to_dict())


class EntryList(list["Entry"]):
    def __init__(self, entries: list["Entry"]) -> None:
        super().__init__(entries)

    @classmethod
    def from_csv_file(cls, file_path: str) -> "EntryList":
        entries: list["Entry"] = []
        with open(file_path, "r") as f:
            reader = csv.DictReader(f)
            for row in reader:
                try:
                    # create
                    entry = Entry.from_dict(row)

                    # check
                    if len(entries) > 0:
                        if entries[-1].date > entry.date:
                            raise Exception(
                                f"date {entry.date} is not greater than {entries[-1].date}"
                            )

                    # append
                    entries.append(entry)
                except Exception as e:
                    raise Exception(f"Index: {reader.line_num}, Row: {row}, Error: {e}")

        return cls(entries)

    @classmethod
    def from_xlsx_file(cls, file_path: str) -> "EntryList":
        entries: list["Entry"] = []
        wb = openpyxl.load_workbook(filename=file_path, read_only=True)
        ws = wb.active
        if ws is None:
            raise Exception(f"{file_path} no active sheet")

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        for row_index, row in enumerate(ws.iter_rows(min_row=2)):
            row_dict = {headers[i]: cell.value for i, cell in enumerate(row)}

            try:
                # create
                entry = Entry.from_dict(row_dict)

                # check
                if len(entries) > 0 and entries[-1].date > entry.date:
                    raise Exception(
                        f"date {entry.date} is not greater than {entries[-1].date}"
                    )

                # append
                entries.append(entry)

            except Exception as e:
                raise Exception(f"Index: {row_index}, Row: {row_dict}, Error: {e}")

        return cls(entries)

    def to_csv_file(self, file_path: str) -> None:
        with open(file_path, "w") as f:
            writer = csv.DictWriter(f, ENTRY_KEYS)
            writer.writeheader()
            for entry in self:
                writer.writerow(entry.to_dict())

    def __str__(self) -> str:
        return "\n".join(str(entry) for entry in self)

    def __repr__(self) -> str:
        return "\n".join(repr(entry) for entry in self)

    def sum(self) -> float:
        total: float = 0
        for entry in self:
            total += SUM_FACTOR[entry.type] * entry.amount
        return total

    @staticmethod
    def concat(last: "EntryList", next: "EntryList") -> "EntryList":
        if len(last) == 0:
            return next
        if len(next) == 0:
            return last
        if last[-1].date > next[0].date:
            raise Exception(f"date {next[0].date} is not greater than {last[-1].date}")
        return EntryList(last + next)

    def filter(self, predicate: Callable[["Entry"], bool]) -> "EntryList":
        return EntryList([entry for entry in self if predicate(entry)])


if __name__ == "__main__":
    entries = EntryList.from_csv_file(get_data_file_path())
    print(f"len(entries): {len(entries)}")
    print(f"entries.sum(): {entries.sum():,.2f}元")
